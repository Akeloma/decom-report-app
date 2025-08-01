
def FLThirtyMth(input_filename):
    import pandas as pd
    from datetime import datetime
    from dateutil.relativedelta import relativedelta
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, PatternFill, Font
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter

    # === Step 1: Load Excel File ===
    # file_path = "manual calculated.xlsx"  # <-- Update this path if needed
    raw_df = pd.read_excel(input_filename, sheet_name="Overall database", header=None)
    
    # === Step 2: Extract headers ===
    header_row_idx = raw_df[raw_df.iloc[:, 0] == "Allianz OE Name"].index[0]
    raw_df.columns = raw_df.iloc[header_row_idx]
    df = raw_df[(header_row_idx + 1):].reset_index(drop=True)
    df.columns = df.columns.map(str).str.strip()

    # === Step 3: Get the latest reporting date ===
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    latest_date = df["Date"].max()
    df = df[df["Date"] == latest_date]

    # === Step 4: Keep relevant columns ===
    df = df[[
        "Allianz OE Name", "IT Component Name", "Release", "IT Component Type",
        "Toxic from Date", "Current Status", "Number of IT Assets"
    ]].copy()

    # === Step 5: Clean data types ===
    df["Toxic from Date"] = pd.to_datetime(df["Toxic from Date"], errors="coerce")
    df["Number of IT Assets"] = pd.to_numeric(df["Number of IT Assets"], errors="coerce")

    # === Step 6: Apply business filters ===
    df = df[
        (df["Current Status"].str.strip() == "Forward Looking Toxic") &
        (df["Number of IT Assets"] > 0)
    ]

    # === Step 7: Filter to assets becoming toxic in next 30 months ===
    end_date = latest_date + relativedelta(months=30)
    df = df[
        (df["Toxic from Date"] >= latest_date) &
        (df["Toxic from Date"] <= end_date)
    ]

    # === Step 8: Add Toxic Year column ===
    df["Toxic Year"] = df["Toxic from Date"].dt.year

    # Define all expected OEs
    all_oe_names = [
        'Allianz China - Holding', 'Allianz China - P&C',
        'Allianz Indonesia', 'Allianz Malaysia',
        'Allianz Philippine - L&H', 'Allianz Singapore',
        'Allianz Sri Lanka', 'Allianz Taiwan - Life',
        'Allianz Thailand'
    ]

    # Define dynamic year range
    year_range = list(range(latest_date.year, (latest_date + relativedelta(months=30)).year + 1))


    # === Step 9: Pivot table builder (dynamic years + Grand Total) ===
    def build_final_table(input_df):
        pivot = input_df.pivot_table(
            index=["Allianz OE Name", "IT Component Name", "Release", "Toxic from Date"],
            columns="Toxic Year",
            values="Number of IT Assets",
            aggfunc="sum",
            fill_value=0
        ).reset_index()

        # Ensure all expected years are present
        for year in year_range:
            if year not in pivot.columns:
                pivot[year] = 0

        # Create rows for missing OEs
        existing_oes = pivot["Allianz OE Name"].unique().tolist()
        for oe in all_oe_names:
            if oe not in existing_oes:
                empty_row = {
                    "Allianz OE Name": oe,
                    "IT Component Name": "-",
                    "Release": "-",
                    "Toxic from Date": pd.NaT,
                    **{year: 0 for year in year_range},
                    "Grand Total": 0
                }
                pivot = pd.concat([pivot, pd.DataFrame([empty_row])], ignore_index=True)

        # Add Grand Total
        pivot["Grand Total"] = pivot[year_range].sum(axis=1)

        return pivot


    # === Step 10: Split by Component Type and create final tables ===
    group_df = df[df["IT Component Type"] == "Group"]
    local_df = df[df["IT Component Type"] != "Group"]

    group_final = build_final_table(group_df)
    local_final = build_final_table(local_df)

    # === Step X: Blank repeated Allianz OE Names, keep only first row ===

    def insert_totals_by_oe(df, year_cols):
        result = []
        for oe_name in df["Allianz OE Name"].unique():
            oe_df = df[df["Allianz OE Name"] == oe_name].copy()
            result.extend(oe_df.to_dict("records"))
            
            total_row = {
                "Allianz OE Name": f"{oe_name} Total",
                "IT Component Name": "",
                "Release": "",
                "Toxic from Date": "",
                **{year: oe_df[year].replace("-", 0).astype(int).sum() for year in year_cols},
                "Grand Total": oe_df["Grand Total"].replace("-", 0).astype(int).sum()
            }

            # Convert 0s back to "-"
            for year in year_cols + ["Grand Total"]:
                if total_row[year] == 0:
                    total_row[year] = "-"

            result.append(total_row)
        return pd.DataFrame(result)


    # NEW STEP: Replace 0s with '-'
    for df in [group_final, local_final]:
        # Identify year columns dynamically in this df
        year_cols = [col for col in df.columns if isinstance(col, int)]
        
        for col in year_cols + ["Grand Total"]:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: "-" if x == 0 else x)

    for df in [group_final, local_final]:
        df["Toxic from Date"] = df["Toxic from Date"].dt.strftime("%d %b %Y")

    #  REMOVE THIS IN CASE
    group_final = insert_totals_by_oe(group_final, year_cols)
    local_final = insert_totals_by_oe(local_final, year_cols)


    # === Step 11: Save to Excel ===
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Group FLT"
    ws2 = wb.create_sheet("Regional_Local FLT")

    def write_to_sheet(ws, df, title):
        ws.sheet_view.showGridLines = False
        ws["A1"] = title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        prev_oe_name = None
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
            is_total_row = isinstance(row[0], str) and row[0].endswith("Total")
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")

                if r_idx == 3:
                    # Header row styling
                    cell.fill = PatternFill(start_color="DDEBF7", fill_type="solid")
                    cell.font = Font(bold=True)

                elif is_total_row:
                    # OE Total row styling
                    cell.fill = PatternFill(start_color="D0CECE", fill_type="solid")
                    cell.font = Font(bold=True)

                elif c_idx == 1:
                    if not is_total_row:
                        if prev_oe_name == value:
                            cell.font = Font(color="FFFFFF")  # white text for duplicates
                        else:
                            cell.font = Font(bold=True)
                            prev_oe_name = value  # update to current OE


        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(str(ws.cell(row=row, column=col_idx).value)) if ws.cell(row=row, column=col_idx).value else 0
                for row in range(3, ws.max_row + 1)
            )
            ws.column_dimensions[col_letter].width = max_len + 2

    write_to_sheet(ws1, group_final, "Group FLT Assets by Year")
    write_to_sheet(ws2, local_final, "Regional/Local FLT Assets by Year")

    # === Step 12: Save Output File ===
    output_file = "ThisFileWorks.xlsx"
    wb.save(output_file)
    print(f"✅ File saved as: {output_file}")
