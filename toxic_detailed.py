import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.styles import Alignment


def main(filename): 

    # === Load Excel File ===
    file_path = "manual calculated.xlsx"
    df = pd.read_excel(file_path, sheet_name="Overall database", skiprows=5)

    # === Clean Column Names ===
    df.columns = df.columns.str.strip()

    # === Read start and end dates from Toxic & FLT Report sheet ===
    wb = load_workbook(file_path, data_only=True)
    ws = wb["Toxic & FLT Report"]

    start_date = pd.to_datetime(ws["G1"].value)
    end_date = pd.to_datetime(ws["G2"].value)


    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df_filtered = df[
        (df["Current Status"] == "Toxic") &
        (df["Date"].between(start_date, end_date)) &
        (df["Number of IT Assets"].notna()) &
        (df["Number of IT Assets"] > 0)
    ].copy()

    df_filtered["Month"] = df_filtered["Date"].dt.normalize()
    df_filtered["Asset ID"] = df_filtered["IT Component Name"] + "|" + df_filtered["Release"] # HEREHERE

    # === Only Compare Start vs End Dates ===
    available_dates = sorted(df_filtered["Date"].dropna().unique())

    # Find first available date >= start_date
    month_start = next((d for d in available_dates if d >= start_date), None)
    # Find first available date <= end_date (most recent before or on end_date)
    month_end = next((d for d in reversed(available_dates) if d <= end_date), None)

    print(f"🧾 Auto-corrected start: {month_start.date() if month_start else 'None'}, end: {month_end.date() if month_end else 'None'}")

    df_compare = df_filtered[df_filtered["Month"].isin([month_start, month_end])]

    # === Group and Aggregate ===
    group_cols = [
        "IT Component Type", "Current Status", "Allianz OE Name", "IT Component Name", "Release", "Month"
    ]
    agg_df = df_compare.groupby(group_cols, as_index=False)["Number of IT Assets"].sum()

    # === Pivot Table for Comparison ===
    pivot_df = agg_df.pivot_table(
        index=["IT Component Type", "Current Status", "Allianz OE Name", "IT Component Name", "Release"],
        columns="Month",
        values="Number of IT Assets",
        fill_value=0
    ).reset_index()

    pivot_df.columns.name = None

    # === Rename Columns (Safely) ===
    available_cols = pivot_df.columns.tolist()

    rename_map = {}
    start_col = None
    end_col = None

    # Rename if available
    for col in available_cols:
        if isinstance(col, pd.Timestamp):
            if col.normalize() == month_start:
                start_col = f"As at {col.day} {col.strftime('%b %Y')}"
                rename_map[col] = start_col
            elif col.normalize() == month_end:
                end_col = f"As at {col.day} {col.strftime('%b %Y')}"
                rename_map[col] = end_col
                
    print(f"✔ Renamed columns: {rename_map}")
    print(f"✔ start_col = {start_col}, end_col = {end_col}")

    pivot_df = pivot_df.rename(columns=rename_map)

    # === Safe Comparison
    if start_col in pivot_df.columns and end_col in pivot_df.columns:
        # Check if both columns are all zero
        if pivot_df[start_col].sum() == 0 and pivot_df[end_col].sum() == 0:
            print(f"⚠️ Both months exist but have only 0 assets. No delta to compute.")
            pivot_df["Delta"] = 0
            pivot_df["Added"] = 0
            pivot_df["Detoxed"] = 0
            pivot_df["Carried Over"] = 0
        else:
            pivot_df["Delta"] = (pivot_df[end_col] - pivot_df[start_col]).abs().fillna(0)
            pivot_df["Added"] = (pivot_df[end_col] - pivot_df[start_col]).clip(lower=0).fillna(0)
            pivot_df["Detoxed"] = (pivot_df[start_col] - pivot_df[end_col]).clip(lower=0).fillna(0)
            pivot_df["Carried Over"] = pivot_df.apply(
                lambda x: x[end_col] if x[end_col] == x[start_col] and x[end_col] != 0 else 0,
                axis=1
            )
            pivot_df["Carried Over"] = pivot_df["Carried Over"].fillna(0)
    else:
        print(f"⚠️ One of the months is missing — skipping delta calculations.")
        pivot_df["Delta"] = 0
        pivot_df["Added"] = 0
        pivot_df["Detoxed"] = 0
        pivot_df["Carried Over"] = 0


    # === Split Group and Local Tables ===
    group_table = pivot_df[pivot_df["IT Component Type"] == "Group"].copy()
    local_table = pivot_df[pivot_df["IT Component Type"] == "Regional/Local"].copy()

    # === Add Total Rows ===
    def add_total_row(df):
        total_row = {
            col: df[col].sum() if df[col].dtype in ['int64', 'float64'] else ''
            for col in df.columns
        }
        total_row["IT Component Type"] = "Total"
        return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

    group_table = add_total_row(group_table)
    local_table = add_total_row(local_table)

    # === Export to Excel - Side by Side in Toxic & FLT Report sheet ===
    wb = load_workbook(file_path)
    ws = wb["Toxic & FLT Report"]

    # Clear previous content from row 15 onward, columns A to M


    # === Add Labels ===
    ws["A49"] = "Group Toxic Details"
    ws["M49"] = "Regional/Local Toxic Details"

    # Write group table starting at A15

    # === Style headers above both tables ===
    ws["A49"] = "Group Toxic Details"
    ws["A49"].fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    ws["A49"].font = Font(bold=True, color="FFFFFF")

    ws["M49"] = "Regional/Local Toxic Details"
    ws["M49"].fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    ws["M49"].font = Font(bold=True, color="FFFFFF")

    for r_idx, row in enumerate(dataframe_to_rows(group_table, index=False, header=True), start=50):
        for c_idx, value in enumerate(row, start=1):  # A = 1
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.value = value

            # Header row
            if r_idx == 50:
                cell.fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
                cell.font = Font(bold=True, color="000000")
            
            # Total row
            if row[0] == "Total":
                if c_idx <= 7:  # For columns A-G
                    cell.fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
                    cell.font = Font(bold=True, color="000000")
                else:  # For columns H-K
                    cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")


    # Write local table starting at M15 (column 13)
    for r_idx, row in enumerate(dataframe_to_rows(local_table, index=False, header=True), start=50):
        for c_idx, value in enumerate(row, start=13):  # M = 13
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.value = value

            # Header row
            if r_idx == 50:
                cell.fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
                cell.font = Font(bold=True, color="000000")
            
            # Total row
            if row[0] == "Total":
                if c_idx <= 19:  # Columns M-S
                    cell.fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
                    cell.font = Font(bold=True, color="000000")
                else:  # Columns T-W
                    cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")

            for col in ws.columns:
                max_length = 0
                column = col[0].column  # Get the column index (1-based)
                column_letter = get_column_letter(column)
            
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            
            
                adjusted_width = max_length + 2  # Add a little padding
                ws.column_dimensions[column_letter].width = adjusted_width

    # === Save the file ===
    wb.save(file_path)
    print("✅ Toxic Detailed Tables Done!")
