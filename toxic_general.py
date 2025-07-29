import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def main(filename):

    # === Load your Excel file ===
    file_path = filename
    df_raw = pd.read_excel(file_path, sheet_name="Overall database", skiprows=5)

    # === Clean and prepare data ===
    df_raw.columns = df_raw.columns.str.strip()
    df = df_raw.copy()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

    # === Read date range from sheet ===
    wb = load_workbook(file_path, data_only=True)
    ws = wb["Toxic & FLT Report"]

    start_date = pd.to_datetime(ws["G1"].value)
    end_date = pd.to_datetime(ws["G2"].value)


    # === Generate month list ===
    month_list = []
    dt = start_date.replace(day=1)
    while dt <= end_date:
        month_list.append(dt.strftime("%b %Y"))
        dt += pd.DateOffset(months=1)

    # === Define OE list ===
    OEs = [
        "Allianz China - Holding", "Allianz China - P&C", "Allianz Indonesia",
        "Allianz Malaysia", "Allianz Philippine - L&H", "Allianz Singapore",
        "Allianz Sri Lanka", "Allianz Taiwan - Life", "Allianz Thailand"
    ]

    # === Filter Toxic data in range with assets > 0 ===
    toxic = df[
        (df["Current Status"] == "Toxic") &
        (df["Number of IT Assets"].fillna(0) > 0) &
        (df["Date"] >= start_date) &
        (df["Date"] <= end_date)
    ].copy()

    toxic["Month"] = toxic["Date"].dt.strftime("%b %Y")

    # === Create monthly totals per OE + ITC type ===
    dictGroup = defaultdict(float)
    dictLocal = defaultdict(float)

    for _, row in toxic.iterrows():
        key = f"{row['Month']}|{row['Allianz OE Name']}"
        val = row["Number of IT Assets"]
        if row["IT Component Type"] == "Group":
            dictGroup[key] += val
        elif row["IT Component Type"] == "Regional/Local":
            dictLocal[key] += val

    # === Calculate MoM changes for each OE ===
    group_rows = []
    local_rows = []

    for oe in OEs:
        groupDelta = groupAdd = groupDetox = groupFinal = 0
        localDelta = localAdd = localDetox = localFinal = 0
        hasLocalChange = False

        for m in range(len(month_list) - 1):
            prev_key = f"{month_list[m]}|{oe}"
            curr_key = f"{month_list[m+1]}|{oe}"

            # Group comparison
            prevG = dictGroup.get(prev_key, 0)
            currG = dictGroup.get(curr_key, 0)

            groupDelta += abs(currG - prevG)
            if currG > prevG:
                groupAdd += currG - prevG
            if currG < prevG:
                groupDetox += prevG - currG
            groupFinal = currG

            # Local comparison
            prevL = dictLocal.get(prev_key, 0)
            currL = dictLocal.get(curr_key, 0)

            if prevL != 0 or currL != 0:
                hasLocalChange = True
                localDelta += abs(currL - prevL)
                if currL > prevL:
                    localAdd += currL - prevL
                if currL < prevL:
                    localDetox += prevL - currL
            localFinal = currL

        # Group result
        group_rows.append({
            "OE": oe,
            "Delta": "-" if groupDelta == 0 else groupDelta,
            "Added": "-" if groupAdd == 0 else groupAdd,
            "Detoxed": "-" if groupDetox == 0 else groupDetox,
            "Carried Over": "-" if groupFinal == 0 else groupFinal,
            f"As at {end_date.strftime('%d %B %Y')}": groupFinal
        })

        # Local result
        if hasLocalChange:
            local_rows.append({
                "OE": oe,
                "Delta": "-" if localDelta == 0 else localDelta,
                "Added": "-" if localAdd == 0 else localAdd,
                "Detoxed": "-" if localDetox == 0 else localDetox,
                "Carried Over": "-" if localFinal == 0 else localFinal,
                f"As at {end_date.strftime('%d %B %Y')}": localFinal
            })
        else:
            local_rows.append({
                "OE": oe,
                "Delta": "-",
                "Added": "-",
                "Detoxed": "-",
                "Carried Over": "-",
                f"As at {end_date.strftime('%d %B %Y')}": 0
            })

    # === Final DataFrames ===
    group_df = pd.DataFrame(group_rows)
    local_df = pd.DataFrame(local_rows)

    # === Load workbook and target sheet ===
    wb = load_workbook("manual calculated.xlsx")
    ws = wb["Toxic & FLT Report"]

    # === Clear existing content (optional) ===
    for row in ws.iter_rows(min_row=4, max_row=15, min_col=1, max_col=13):
        for cell in row:
            cell.value = None

    # === Styling setup ===
    header_fill = PatternFill(start_color="122B54", end_color="122B54", fill_type="solid")
    section_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    total_fill = PatternFill(start_color="122B54", end_color="122B54", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")


    def write_table(sheet, df, start_row, start_col, section_title):
        # Write section title
        title_cell = sheet.cell(row=start_row, column=start_col, value=section_title)
        title_cell.fill = section_fill
        title_cell.font = white_font
        title_cell.alignment = center_align

        # Write headers and data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            for c_idx, value in enumerate(row):
                cell = sheet.cell(row=start_row + 1 + r_idx, column=start_col + c_idx, value=value)
                cell.alignment = center_align
                if r_idx == 0:  # Header row
                    cell.fill = header_fill
                    cell.font = white_font

        # Add totals row (after data)
        data_start = start_row + 2
        data_end = data_start + len(df) - 1
        total_row_idx = data_end + 1
        sheet.cell(row=total_row_idx, column=start_col, value="Total")
        sheet.cell(row=total_row_idx, column=start_col).fill = total_fill
        sheet.cell(row=total_row_idx, column=start_col).font = white_font
        sheet.cell(row=total_row_idx, column=start_col).alignment = center_align

        for i in range(1, df.shape[1]):
            col_letter = get_column_letter(start_col + i)
            formula = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
            cell = sheet.cell(row=total_row_idx, column=start_col + i, value=formula)
            cell.fill = total_fill
            cell.font = white_font
            cell.alignment = center_align

        # Auto-fit columns
        for col_idx in range(start_col, start_col + df.shape[1]):
            max_len = 0
            for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx, min_row=start_row, max_row=total_row_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


    write_table(ws, group_df, start_row=4, start_col=1, section_title="Group Toxic General")
    write_table(ws, local_df, start_row=4, start_col=8, section_title="Regional/Local Toxic General")


    # === Save workbook ===
    wb.save("manual calculated.xlsx")
    print("✅ General Toxic Tables Done!")
