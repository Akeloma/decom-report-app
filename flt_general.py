import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

def main():

    # === File paths ===
    file_path = "manual calculated.xlsx"
    sheet_name = "Overall database"

    # === Load data ===
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=5)
    df.columns = df.columns.str.strip()
    all_oes = sorted(set(df['Allianz OE Name'].dropna()) - {'Allianz Laos'})


    # === Read start and end dates from Toxic & FLT Report sheet ===
    wb = load_workbook(file_path, data_only=True)
    ws = wb["Toxic & FLT Report"]

    start_date = pd.to_datetime(ws["G1"].value)
    end_date = pd.to_datetime(ws["G2"].value)

    # === Filter
    flt_df = df[
        (df['Current Status'] == 'Forward Looking Toxic') &
        (df['Toxic from Date'].astype(str).str.contains('2025')) &
        (df['Number of IT Assets'] > 0) &
        (df['Date'].notna()) &
        (df['Date'] >= start_date) &
        (df['Date'] <= end_date)
    ].copy()

    flt_df.sort_values(by='Date', inplace=True)
    unique_dates = flt_df['Date'].drop_duplicates().sort_values().to_list()
    month_pairs = [(unique_dates[i], unique_dates[i+1]) for i in range(len(unique_dates)-1)]

    summary_rows = []
    for prev_date, curr_date in month_pairs:
        prev_df = flt_df[flt_df['Date'] == prev_date]
        curr_df = flt_df[flt_df['Date'] == curr_date]
        prev_df['Key'] = prev_df['Allianz OE Name'].astype(str) + '|' + prev_df['IT Component Name'].astype(str) + '|' + prev_df['IT Component Type'].astype(str) + '|' + prev_df['Release'].astype(str)
        curr_df['Key'] = curr_df['Allianz OE Name'].astype(str) + '|' + curr_df['IT Component Name'].astype(str) + '|' + curr_df['IT Component Type'].astype(str) + '|' + curr_df['Release'].astype(str)

        prev_dict = dict(zip(prev_df['Key'], prev_df['Number of IT Assets']))
        curr_dict = dict(zip(curr_df['Key'], curr_df['Number of IT Assets']))
        type_dict = dict(zip(curr_df['Key'], curr_df['IT Component Type']))
        oe_dict = dict(zip(curr_df['Key'], curr_df['Allianz OE Name']))

        all_keys = set(prev_dict.keys()).union(set(curr_dict.keys()))
        for key in all_keys:
            oe = oe_dict.get(key) or key.split('|')[0]
            typ = type_dict.get(key) or key.split('|')[2]
            prev_val = prev_dict.get(key, 0)
            curr_val = curr_dict.get(key, 0)
            added = max(0, curr_val - prev_val)
            detoxed = max(0, prev_val - curr_val)
            delta = abs(added - detoxed)
            carried = prev_val - detoxed + added

            summary_rows.append({
                'OE': oe,
                'Asset Type': typ,
                'Delta': delta,
                'Added': added,
                'Detoxed': detoxed,
                f'Total {curr_date.strftime("%d-%b")}': curr_val
            })

    summary_df = pd.DataFrame(summary_rows)
    final_summary = summary_df.groupby(['OE', 'Asset Type'], as_index=False).sum()
    final_summary['Delta'] = abs(final_summary['Added'] - final_summary['Detoxed'])
    final_summary['Carried Over'] = final_summary[
        [col for col in final_summary.columns if col.startswith('Total ')]
    ].max(axis=1) - final_summary['Added'] + final_summary['Detoxed']
    final_summary = final_summary[['OE', 'Asset Type', 'Added', 'Detoxed', 'Delta', 'Carried Over'] +
                                [col for col in final_summary.columns if col.startswith('Total ')]]

    totals_row = final_summary.drop(columns=['OE', 'Asset Type']).sum(numeric_only=True)
    totals_row['OE'] = 'Total'
    totals_row['Asset Type'] = ''
    final_summary = pd.concat([final_summary, totals_row.to_frame().T], ignore_index=True)

    # === Group / Local tables
    group_tbl = final_summary[final_summary['Asset Type'] == 'Group'].copy().set_index('OE').reindex(all_oes, fill_value=0).reset_index()
    group_tbl['Asset Type'] = 'Group'
    local_tbl = final_summary[final_summary['Asset Type'] == 'Regional/Local'].copy().set_index('OE').reindex(all_oes, fill_value=0).reset_index()
    local_tbl['Asset Type'] = 'Regional/Local'

    for tbl in [group_tbl, local_tbl]:
        tbl.fillna(0, inplace=True)
        total_row = tbl.drop(columns=['OE', 'Asset Type']).sum(numeric_only=False)
        total_row['OE'] = 'Total'
        total_row['Asset Type'] = ''
        tbl.loc[len(tbl.index)] = total_row

    # === Format & Write
    def format_table(ws, start_row, table_df, title):
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="122B54", end_color="122B54", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=10)
        title_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        total_fill = PatternFill(start_color="122B54", end_color="122B54", fill_type="solid")
        total_font = Font(bold=True, color="FFFFFF", size=10)
        align_center = Alignment(horizontal="center", vertical="center")

        title_cell = ws.cell(row=start_row, column=1, value=title)
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = align_center
        #ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(table_df.columns))

        header_row = start_row + 1
        for col_idx, col_name in enumerate(table_df.columns, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        for row_idx, row_data in enumerate(table_df.values, header_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(size=10)
                cell.alignment = align_center

        total_row_idx = header_row + len(table_df)
        for col_idx in range(1, len(table_df.columns) + 1):
            letter = get_column_letter(col_idx)
            max_length = len(str(table_df.columns[col_idx - 1]))  # header length

        # Check all cell values in that column
        for row in table_df.itertuples(index=False):
            try:
                value_len = len(str(row[col_idx - 1]))
                max_length = max(max_length, value_len)
            except IndexError:
                pass

        ws.column_dimensions[letter].width = max_length + 2

        for col_idx in range(1, len(table_df.columns) + 1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = min(max(len(str(table_df.columns[col_idx - 1])) + 2, 10), 25)

        return total_row_idx + 2

    # === Export
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb['Toxic & FLT Report']
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Toxic & FLT Report'

    row = 20
    row = format_table(ws, row, group_tbl, "Group FLT General")
    row = format_table(ws, row, local_tbl, "Regional/Local FLT General")

    wb.save(file_path)
    print("General FLT Tables Done!")
