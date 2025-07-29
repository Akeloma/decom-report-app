import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from collections import defaultdict


def main(filename):

    # === CONFIGURATION ===
    file_path = filename
    wb = load_workbook(file_path, data_only=True)
    ws = wb["Toxic & FLT Report"]

    start_date = pd.to_datetime(ws["G1"].value)
    end_date = pd.to_datetime(ws["G2"].value)



    # === LOAD & CLEAN DATA ===
    df = pd.read_excel(file_path, sheet_name="Overall database", skiprows=5)
    df.columns = df.columns.str.strip()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

    df = df[
        (df["Date"] >= start_date) &
        (df["Date"] <= end_date) &
        (df["Current Status"] == "Forward Looking Toxic") &
        (df["Allianz OE Name"] != "Allianz Laos") &
        (df["Toxic from Date"].astype(str).str.contains("2025", na=False)) &
        (df["Number of IT Assets"] > 0)
    ]

    df["Key"] = (
        df["IT Component Type"].astype(str) + "|" +
        df["Allianz OE Name"].astype(str) + "|" +
        df["IT Component Name"].astype(str) + "|" +
        df["Release"].astype(str)
    )
    df["Month"] = df["Date"].dt.to_period("M")

    # === PROCESS MONTH PAIRS ===
    month_list = sorted(df["Month"].unique())
    summary = defaultdict(lambda: {"Added": 0, "Detoxed": 0, "Delta": 0, "Carried": 0, "Details": None})

    for i in range(len(month_list) - 1):
        m1, m2 = month_list[i], month_list[i + 1]
        df1 = df[df["Month"] == m1]
        df2 = df[df["Month"] == m2]

        dict1 = df1.groupby("Key")["Number of IT Assets"].sum().to_dict()
        dict2 = df2.groupby("Key")["Number of IT Assets"].sum().to_dict()

        for key in set(dict1).union(dict2):
            val1, val2 = dict1.get(key, 0), dict2.get(key, 0)
            delta = abs(val2 - val1)
            added = val2 if val1 == 0 and val2 > 0 else 0
            detoxed = val1 if val1 > 0 and val2 == 0 else 0
            carried = min(val1, val2) if val1 > 0 and val2 > 0 else 0

            summary[key]["Added"] += added
            summary[key]["Detoxed"] += detoxed
            summary[key]["Carried"] += carried
            summary[key]["Delta"] += delta
            summary[key]["Current"] = val2  # asset count in end month
            summary[key]["Previous"] = val1  # asset count in start month


            if summary[key]["Details"] is None:
                parts = key.split("|")
                asset_date = df[df["Key"] == key]["Date"].iloc[0] if not df[df["Key"] == key].empty else None
                formatted_date = f"{asset_date.day} {asset_date.strftime('%b %Y')}" if pd.notnull(asset_date) else ""
                
                summary[key]["Details"] = {
                    "IT Component Type": parts[0],
                    "Current Status": "Forward Looking Toxic",
                    "Date": formatted_date,
                    "Allianz OE Name": parts[1],
                    "IT Component Name": parts[2],
                    "Release": parts[3]
                }



    # === BUILD OUTPUT DATAFRAMES ===
    rows = []
    for key, val in summary.items():
        if val["Added"] == 0 and val["Detoxed"] == 0 and val["Carried"] == 0 and val["Delta"] == 0:
            continue
        details = val["Details"]

        # Pull FLT values at start_date and end_date
        start_value = df[
            (df["Key"] == key) & (df["Date"] == start_date)
        ]["Number of IT Assets"].sum()

        end_value = df[
            (df["Key"] == key) & (df["Date"] == end_date)
        ]["Number of IT Assets"].sum()

        start_label = f"As of ({start_date.strftime('%-d %b %Y')})"
        end_label = f"As of ({end_date.strftime('%-d %b %Y')})"

        rows.append({
            **details,
            start_label: int(start_value),
            end_label: int(end_value),
            "Delta": val["Delta"],
            "Added": val["Added"],
            "Detoxed": val["Detoxed"],
            "Carried Over": val["Carried"]
        })



    df_result = pd.DataFrame(rows)
    group_df = df_result[df_result["IT Component Type"].str.upper().str.strip() == "GROUP"]
    regional_df = df_result[df_result["IT Component Type"].str.upper().str.strip() == "REGIONAL/LOCAL"]

    # === WRITE TO EXCEL ===
    wb = load_workbook(file_path)
    ws = wb["Toxic & FLT Report"]

    # Clear area
    # for row in ws.iter_rows(min_row=25, max_row=150, min_col=1, max_col=25):
    #     for cell in row:
    #         cell.value = None

    # Title rows
    ws["A64"] = "Group FLT Details"
    ws["O64"] = "Regional/Local FLT Details"
    for cell in [ws["A64"], ws["O64"]]:
        cell.fill = PatternFill("solid", fgColor="00B0F0")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Headers formatting
    header_fill = PatternFill("solid", fgColor="E4DFEC")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    def write_table(df, start_col):
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=65):
            for c_idx, val in enumerate(row, start=start_col):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = center_align
                if r_idx == 65:
                    cell.fill = header_fill
                    cell.font = header_font
        # Total row
        total_row = 65 + len(df) + 1
        ws.cell(row=total_row, column=start_col, value="Total").alignment = center_align

        # Fill the entire Total row with header styling
        num_cols = len(df.columns)
        for col_idx in range(start_col, start_col + num_cols):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.alignment = center_align
            cell.fill = header_fill
            cell.font = header_font
        
        from openpyxl.utils import get_column_letter

        # Write the "Total" label
        ws.cell(row=total_row, column=start_col, value="Total")

        # Dynamically sum all numeric columns (except text ones like "IT Component Name")
        for i, col_name in enumerate(df.columns, start=start_col):
            if df[col_name].dtype in [int, float]:
                col_letter = get_column_letter(i)
                formula = f"=SUM({col_letter}66:{col_letter}{total_row-1})"
                cell = ws.cell(row=total_row, column=i)
                cell.value = formula
                cell.alignment = center_align
                cell.fill = header_fill
                cell.font = header_font

    write_table(group_df, 1)
    write_table(regional_df, 15)

    wb.save(file_path)
    print("✅ FLT Detailed Tables Done!")
    print("Group rows:", len(group_df))
    print("Regional rows:", len(regional_df))
