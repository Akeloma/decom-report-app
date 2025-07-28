import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side


def main():

    # === Load Excel File ===
    file_path = "manual calculated.xlsx"  # Update if needed
    df_raw = pd.read_excel(file_path, sheet_name="Overall database", skiprows=5)
    df_raw.columns = df_raw.columns.str.strip()
    df_raw["Date"] = pd.to_datetime(df_raw["Date"], errors="coerce")

    # === Get Start/End Dates ===
    wb = load_workbook(file_path, data_only=True)
    ws = wb["Toxic & FLT Report"]
    start_date = pd.to_datetime(ws["G1"].value)
    end_date = pd.to_datetime(ws["G2"].value)

    # === Define OE List ===
    OEs = [
        "Allianz China - Holding", "Allianz China - P&C", "Allianz Indonesia",
        "Allianz Malaysia", "Allianz Philippine - L&H", "Allianz Singapore",
        "Allianz Sri Lanka", "Allianz Taiwan - Life", "Allianz Thailand"
    ]

    # === Create Month List ===
    month_list = []
    dt = start_date.replace(day=1)
    while dt <= end_date:
        month_list.append(dt.strftime("%b %Y"))
        dt += pd.DateOffset(months=1)

    # === Toxic Section ===
    toxic_df = df_raw[
        (df_raw["Current Status"] == "Toxic") &
        (df_raw["Number of IT Assets"].fillna(0) > 0) &
        (df_raw["Date"].between(start_date, end_date))
    ].copy()
    toxic_df["Month"] = toxic_df["Date"].dt.strftime("%b %Y")

    dictGroup, dictLocal = defaultdict(float), defaultdict(float)

    for _, row in toxic_df.iterrows():
        key = f"{row['Month']}|{row['Allianz OE Name']}"
        val = row["Number of IT Assets"]
        if row["IT Component Type"] == "Group":
            dictGroup[key] += val
        elif row["IT Component Type"] == "Regional/Local":
            dictLocal[key] += val

    group_rows, local_rows = [], []

    for oe in OEs:
        groupDetox = localDetox = 0
        groupFinal = localFinal = 0
        hasLocalChange = False

        for m in range(len(month_list) - 1):
            prev_key = f"{month_list[m]}|{oe}"
            curr_key = f"{month_list[m+1]}|{oe}"
            prevG = dictGroup.get(prev_key, 0)
            currG = dictGroup.get(curr_key, 0)
            prevL = dictLocal.get(prev_key, 0)
            currL = dictLocal.get(curr_key, 0)

            if currG < prevG:
                groupDetox += prevG - currG
            groupFinal = currG

            if currL != 0 or prevL != 0:
                hasLocalChange = True
                if currL < prevL:
                    localDetox += prevL - currL
            localFinal = currL

        group_rows.append({
            "OE": oe,
            "2025 YTD Detoxed": -groupDetox if groupDetox > 0 else 0,
            "2025 Toxic": groupFinal
        })

        local_rows.append({
            "OE": oe,
            "2025 YTD Detoxed": abs(localDetox) if hasLocalChange and localDetox > 0 else 0,
            "2025 Toxic": localFinal
        })

    group_toxic_df = pd.DataFrame(group_rows)
    local_toxic_df = pd.DataFrame(local_rows)

    # === FLT Section ===
    flt_df = df_raw[
        (df_raw["Current Status"] == "Forward Looking Toxic") &
        (df_raw["Toxic from Date"].astype(str).str.contains('2025')) &
        (df_raw["Number of IT Assets"] > 0) &
        (df_raw["Date"].notna()) &
        (df_raw["Date"].between(start_date, end_date))
    ].copy()
    flt_df.sort_values(by='Date', inplace=True)

    unique_dates = flt_df['Date'].drop_duplicates().sort_values().to_list()
    month_pairs = [(unique_dates[i], unique_dates[i+1]) for i in range(len(unique_dates)-1)]

    summary_rows = []
    for prev_date, curr_date in month_pairs:
        prev_df = flt_df[flt_df['Date'] == prev_date]
        curr_df = flt_df[flt_df['Date'] == curr_date]

        prev_df['Key'] = prev_df['Allianz OE Name'] + '|' + prev_df['IT Component Name'] + '|' + prev_df['IT Component Type'] + '|' + prev_df['Release'].astype(str)
        curr_df['Key'] = curr_df['Allianz OE Name'] + '|' + curr_df['IT Component Name'] + '|' + curr_df['IT Component Type'] + '|' + curr_df['Release'].astype(str)

        prev_dict = dict(zip(prev_df['Key'], prev_df['Number of IT Assets']))
        curr_dict = dict(zip(curr_df['Key'], curr_df['Number of IT Assets']))
        type_dict = dict(zip(curr_df['Key'], curr_df['IT Component Type']))
        oe_dict = dict(zip(curr_df['Key'], curr_df['Allianz OE Name']))

        all_keys = set(prev_dict.keys()).union(curr_dict.keys())
        for key in all_keys:
            oe = oe_dict.get(key, key.split('|')[0])
            typ = type_dict.get(key, key.split('|')[2])
            detoxed = max(0, prev_dict.get(key, 0) - curr_dict.get(key, 0))
            summary_rows.append({"OE": oe, "Asset Type": typ, "2025 FLT Detoxed": detoxed})

    flt_detox_df = pd.DataFrame(summary_rows).groupby(['OE', 'Asset Type'], as_index=False).sum()

    # === Latest FLT Snapshot ===
    latest_date = flt_df['Date'].max()
    latest_snapshot = flt_df[flt_df['Date'] == latest_date]
    latest_group = latest_snapshot.groupby(['Allianz OE Name', 'IT Component Type'])['Number of IT Assets'].sum().reset_index()
    latest_group.rename(columns={
        'Allianz OE Name': 'OE',
        'IT Component Type': 'Asset Type',
        'Number of IT Assets': '2025 FLT'
    }, inplace=True)

    # === Combine Detoxed + FLT ===
    flt_final = pd.merge(flt_detox_df, latest_group, on=['OE', 'Asset Type'], how='outer').fillna(0)

    # === Final Merge Toxic + FLT ===
    group_flt = flt_final[flt_final['Asset Type'] == 'Group'].drop(columns='Asset Type')
    local_flt = flt_final[flt_final['Asset Type'] == 'Regional/Local'].drop(columns='Asset Type')

    group_summary = pd.merge(group_toxic_df, group_flt, on='OE', how='left').fillna(0)
    local_summary = pd.merge(local_toxic_df, local_flt, on='OE', how='left').fillna(0)

    group_summary.insert(1, "Year", current_year)
    local_summary.insert(1, "Year", current_year)


    group_summary["2025 Total (Toxic + FLT)"] = group_summary["2025 Toxic"] + group_summary["2025 FLT"]
    local_summary["2025 Total (Toxic + FLT)"] = local_summary["2025 Toxic"] + local_summary["2025 FLT"]

    # === Optional: Format Columns as Int ===
    cols = ["2025 YTD Detoxed", "2025 Toxic", "2025 FLT", "2025 Total (Toxic + FLT)"]
    for col in cols:
        group_summary[col] = group_summary[col].astype(int)
        local_summary[col] = local_summary[col].astype(int)

    # === Show Final Tables ===
    # print("📘 Group Table:\n", group_summary)
    # print("\n📗 Regional/Local Table:\n", local_summary)

    #------------------------------------------------------------------------------------------------
    # === Define Quarters Function ===
    def get_quarter_with_year(date):
        if pd.isna(date):
            return None
        year = date.year
        if date.month in [1, 2, 3]:
            quarter = 'Q1'
        elif date.month in [4, 5, 6]:
            quarter = 'Q2'
        elif date.month in [7, 8, 9]:
            quarter = 'Q3'
        elif date.month in [10, 11, 12]:
            quarter = 'Q4'
        else:
            return None
    
        if year == 2025:
            return quarter
        else:
            return f'{quarter}_{year}'


    # === Get Latest Snapshot for Toxic ===
    latest_toxic_date = toxic_df['Date'].max()
    latest_toxic = toxic_df[toxic_df['Date'] == latest_toxic_date].copy()
    latest_toxic['Planned Completion Date'] = pd.to_datetime(latest_toxic['Planned Completion Date'], errors='coerce')
    latest_toxic['Quarter'] = latest_toxic['Planned Completion Date'].apply(get_quarter_with_year)
    latest_toxic['Quarter'].fillna('Unknown', inplace=True)
    latest_toxic['Data Type'] = 'Toxic'

    # === Get Latest Snapshot for FLT ===
    latest_flt_date = flt_df['Date'].max()
    latest_flt = flt_df[flt_df['Date'] == latest_flt_date].copy()
    latest_flt['Planned Completion Date'] = pd.to_datetime(latest_flt['Planned Completion Date'], errors='coerce')
    latest_flt['Quarter'] = latest_flt['Planned Completion Date'].apply(get_quarter_with_year)
    latest_flt['Quarter'].fillna('Unknown', inplace=True)
    latest_flt['Data Type'] = 'FLT'

    # === Combine Toxic + FLT Quarters ===
    combined_df = pd.concat([latest_toxic, latest_flt])

    # === Summarize by OE, Asset Type, Quarter, and Data Type ===
    quarter_summary = combined_df.groupby(
        ['Allianz OE Name', 'IT Component Type', 'Data Type', 'Quarter']
    )['Number of IT Assets'].sum().reset_index()

    #----------------------------------------------------------------------------------

    # === Pivot into Q1–Q4 columns ===
    quarter_pivot = quarter_summary.pivot_table(
        index=['Allianz OE Name', 'IT Component Type'],
        columns='Quarter',
        values='Number of IT Assets',
        aggfunc='sum',
        fill_value=0
    ).reset_index()

    # Ensure Q1–Q4 columns exist
    # Ensure all quarter columns exist in pivot (including dynamic ones like Q1_2026)
    all_quarters = quarter_summary['Quarter'].dropna().unique()
    for q in all_quarters:
        if q not in quarter_pivot.columns:
            quarter_pivot[q] = 0



    # Rename for consistency
    quarter_pivot.rename(columns={'Allianz OE Name': 'OE'}, inplace=True)

    # Separate Group and Local
    group_quarters = quarter_pivot[quarter_pivot['IT Component Type'] == 'Group'].drop(columns='IT Component Type')
    local_quarters = quarter_pivot[quarter_pivot['IT Component Type'] == 'Regional/Local'].drop(columns='IT Component Type')

    # === Add Total Row Function ===
    def add_total_row(df):
        quarter_cols = [col for col in df.columns if col.startswith('Q')]
        total = df[quarter_cols].sum()
        total_row = {'OE': 'Total'}
        total_row.update(total.to_dict())
        return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

    # group_tbl = add_total_row(group_tbl)  #UNCOMMENT IF IT DOESNT WORK OUT
    # local_tbl = add_total_row(local_tbl)

    group_tbl = add_total_row(group_quarters)
    local_tbl = add_total_row(local_quarters)

    #------------------------------------------------final output below-------------------

    current_year = pd.Timestamp.now().year

    # === Merge Quarters into Final Summary ===
    group_final = pd.merge(group_summary, group_tbl, on="OE", how="left").fillna(0)
    local_final = pd.merge(local_summary, local_tbl, on="OE", how="left").fillna(0)

    # ✅ Remove "Data Type" AFTER the merge
    group_final.drop(columns=['Data Type'], errors='ignore', inplace=True)
    local_final.drop(columns=['Data Type'], errors='ignore', inplace=True)


    # === Merge FLT detox into YTD Detoxed, then drop FLT Detoxed ===
    group_final["2025 YTD Detoxed"] = group_final["2025 YTD Detoxed"] + group_final["2025 FLT Detoxed"]
    group_final.drop(columns=["2025 FLT Detoxed"], inplace=True)

    local_final["2025 YTD Detoxed"] = local_final["2025 YTD Detoxed"] + local_final["2025 FLT Detoxed"]
    local_final.drop(columns=["2025 FLT Detoxed"], inplace=True)


    # === Reorder Q1–Q4 columns to be in proper order
    import re

    # === Reorder ALL Quarter Columns (Q1, Q2, ..., Q1_2026, Q2_2026, ...)
    def sort_quarters(cols):
        def quarter_key(col):
            match = re.match(r"Q([1-4])(?:_(\d{4}))?", col)
            if match:
                q_num = int(match.group(1))
                year = int(match.group(2)) if match.group(2) else 2025  # Default to 2025 if no year
                return (year, q_num)
            return (9999, 9)  # Put anything non-quarter at the end
        return sorted([c for c in cols if c.startswith("Q")], key=quarter_key)
    
    # Get ordered list of quarters for each table
    group_quarter_cols = sort_quarters(group_final.columns)
    local_quarter_cols = sort_quarters(local_final.columns)
    
    # Reorder columns
    group_non_quarter_cols = [c for c in group_final.columns if c not in group_quarter_cols]
    group_final = group_final[group_non_quarter_cols + group_quarter_cols]
    
    local_non_quarter_cols = [c for c in local_final.columns if c not in local_quarter_cols]
    local_final = local_final[local_non_quarter_cols + local_quarter_cols]

    # === Add Total Row to group_final ===
    group_total = group_final.drop(columns=["Year"]).select_dtypes(include='number').sum()
    group_total["OE"] = "Total"
    group_total["Year"] = ""  # Keep Year column empty for total row
    group_final = pd.concat([group_final, pd.DataFrame([group_total])], ignore_index=True)


    # === Add Total Row to local_final ===
    local_total = local_final.drop(columns=["Year"]).select_dtypes(include='number').sum()
    local_total["OE"] = "Total"
    local_total["Year"] = ""
    local_final = pd.concat([local_final, pd.DataFrame([local_total])], ignore_index=True)

    # Keep only required columns
    keep_cols = ["OE", "Year", "2025 YTD Detoxed", "2025 Toxic", "2025 FLT"]
    group_final = group_final[keep_cols]
    local_final = local_final[keep_cols]


    # === Final Output ===
    print("📘 FINAL Group Table with Quarters:\n", group_final)
    print("\n📗 FINAL Local Table with Quarters:\n", local_final)


    # === Create a new Excel workbook ===
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Group Summary"
    ws2 = wb.create_sheet("Local Summary")

    # === Define thin black border style ===
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    # === Add merged header row for Group Sheet ===
    ws1.merge_cells('A1:I1')
    header_cell = ws1['A1']
    header_cell.value = "Group : Current and Forward Looking Toxic (FLT)"
    header_cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_cell.font = Font(bold=True, color="FFFFFF")
    header_cell.alignment = Alignment(horizontal="center", vertical="center")


    # === Write group_final to Sheet 1 ===
    for r_idx, row in enumerate(dataframe_to_rows(group_final, index=False, header=True), start=2):
        for c_idx, value in enumerate(row, start=1):
            col_letter = group_final.columns[c_idx - 1]

            cell = ws1.cell(row=r_idx, column=c_idx, value=value)

            is_header = r_idx == 2  # row 2 is always header now

            # Borders for everyone
            cell.border = thin_border

            # Apply styles
            if is_header:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            elif col_letter == "2025 YTD Detoxed":
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            elif col_letter == "2025 Total (Toxic + FLT)":
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

            cell.alignment = Alignment(horizontal="center")

    # === Remove fill from Row 1 and Row 12 (Group sheet) from column F onward
    for col in range(6, ws1.max_column + 1):  # F onwards = col 6
        ws1.cell(row=1, column=col).fill = PatternFill(fill_type=None)
        ws1.cell(row=12, column=col).fill = PatternFill(fill_type=None)



        # === Manually style last row in Group sheet ===
    last_row = ws1.max_row
    for col in range(1, ws1.max_column + 1):
        cell = ws1.cell(row=last_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # === Add merged header row for Local Sheet ===
    ws2.merge_cells('A1:I1')
    header_cell = ws2['A1']
    header_cell.value = "Local : Current and Forward Looking Toxic (FLT)"
    header_cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_cell.font = Font(bold=True, color="FFFFFF")
    header_cell.alignment = Alignment(horizontal="center", vertical="center")


    # === Write local_final to Sheet 2 ===
    for r_idx, row in enumerate(dataframe_to_rows(local_final, index=False, header=True), start=2):
        for c_idx, value in enumerate(row, start=1):
            col_letter = local_final.columns[c_idx - 1]

            cell = ws2.cell(row=r_idx, column=c_idx, value=value)

            is_header = r_idx == 2

            # Borders for everyone
            cell.border = thin_border

            # Apply styles
            if is_header:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            elif col_letter == "2025 YTD Detoxed":
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            elif col_letter == "2025 Total (Toxic + FLT)":
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

            cell.alignment = Alignment(horizontal="center")

    # === Remove fill from Row 1 and Row 12 (Local sheet) from column F onward
    for col in range(6, ws2.max_column + 1):
        ws2.cell(row=1, column=col).fill = PatternFill(fill_type=None)
        ws2.cell(row=12, column=col).fill = PatternFill(fill_type=None)


        # === Manually style last row in Local sheet ===
    last_row = ws2.max_row
    for col in range(1, ws2.max_column + 1):
        cell = ws2.cell(row=last_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # === Autofit column widths ===
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

    # === Save the file ===
    wb.save("Toxic&FLT_Tables.xlsx")  # You can change filename as needed

if __name__ == "__main__":
    main()
