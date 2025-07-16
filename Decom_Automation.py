import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Alignment


# Decom_Automation.py
def main():
    try:
        # === CONFIGURATION ===
        file_path = "Decom.xlsx"
        sheet_name = "Raw Data"
        sheet_name2 = "2025P PD24 Decom plan"
        output_sheet = "Decom Dashboard"

        # Hardcoded Decom Plan (PD24)
        #### 2/07 EDITTED To Autoamted Decom Plan(PD24)

        decom_plan={'Allianz China - Holding (CNLH)': 0,
            'Allianz China - P&C': 0,
            'Allianz Indonesia (ID)': 0,
            'Allianz Malaysia (MY)': 0,
            'Allianz Philippine - L&H (PH)': 0,
            'Allianz Singapore (AS)': 0,
            'Allianz Sri Lanka (LK)': 0,
            'Allianz Taiwan - Life (TWL)': 0,
            'Allianz Thailand (TH)': 0}
        df1 = pd.read_excel(file_path, sheet_name=sheet_name2,
                header=2, nrows=10, usecols="B:C")

        for _, row in df1.iterrows():
            key = row['Oes']       # key in dictionary
            value = row['2025 Decom Plan (PD24)']     # value to update/add
            if key in decom_plan:
                decom_plan[key] += value   # add value to existing total

        # === LOAD & FILTER RAW DATA ===
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=1)
        df = df[['OE Name', 'Forecast End Date', 'Phase']].copy()
        df.columns = ['OE', 'Date', 'Status']
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df[df['Date'].dt.year == 2025].copy()
        df['Quarter'] = df['Date'].dt.quarter.map(lambda q: f"Q{q}") # THIS IS CORRECT OUTPUT --> CORRECT FILTER 

        # === TABLE STRUCTURE ===
        oes = list(decom_plan.keys())
        output = pd.DataFrame(index=oes)
        output["2025 Decom Plan (PD24)"] = output.index.map(decom_plan.get)

        # Completed YTD
        output["Completed YTD"] = df[df['Status'].str.lower() == "completed"].groupby("OE").size()
        output["Completed YTD"] = output["Completed YTD"].fillna(0).astype(int) # COMPLETED IS CORRECT NUMBERS 

        # In Progress (exclude completed + descoped)
        mask_in_progress = ~df["Status"].str.lower().isin(["completed", "descoped"])
        df_in_progress = df[mask_in_progress]

        output["In Progress"] = df_in_progress.groupby("OE").size()
        output["In Progress"] = output["In Progress"].fillna(0).astype(int) # IN PROGRESS IS CORRECT NUMBERS

        # Quarter-wise counts (from in-progress only)
        for q in ['Q1', 'Q2', 'Q3', 'Q4']:
            output[q] = df_in_progress[df_in_progress['Quarter'] == q].groupby('OE').size() #CORRECT NUMBERS DISPLAYED
            

        # Fill NaNs for missing quarter-OEs
        output[['Q1', 'Q2', 'Q3', 'Q4']] = output[['Q1', 'Q2', 'Q3', 'Q4']].fillna(0).astype(int)

        # Grand Total by quarter columns
        output["Grand Total"] = output[['Q1', 'Q2', 'Q3', 'Q4']].sum(axis=1)

        # YTD FC = Completed YTD + In Progress
        output["2025 YTD FC"] = output["Completed YTD"] + output["In Progress"]

        # Add Grand Total row
        # if "Grand Total" in output.index:
        #     output = output.drop("Grand Total")

        # Now safely compute grand totals
        grand_total = output.sum(numeric_only=True)
        grand_total["2025 Decom Plan (PD24)"] = sum(decom_plan.values())  # hardcoded sum = 12

        # Append grand total row
        output.loc["Grand Total"] = grand_total

        # === EXPORT TO EXCEL IN TABLE FORMAT ===
        wb = load_workbook(file_path)
        if output_sheet in wb.sheetnames:
            del wb[output_sheet]
        ws = wb.create_sheet(output_sheet)

        # Title Header
        ws["E1"] = "Planned Decom Timeline"
        ws["E1"].font = Font(bold=True)
        ws.merge_cells("E1:I1")

        # Column Headers
        headers = [
            "OE", "2025 Decom Plan (PD24)", "Completed YTD", "In Progress",
            "Q1", "Q2", "Q3", "Q4", "Grand Total", "2025 YTD FC"
        ]
        ws.append(headers)

        # Write Data
        for idx, row in output.reset_index().iterrows():
            ws.append(list(row))

        # Header Colors
        header_fill = PatternFill(start_color='FF70AD47', end_color='FF70AD47', fill_type='solid')  # Green
        data_fill = PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid') # Light Green

        # Initialise borders
        thin_border = Border(bottom=Side(style='thin', color='000000'))

        # Color in headers
        for col in range(1, len(headers) + 1):
            for row in range(1, 3):
                cell = ws[f"{get_column_letter(col)}{row}"]
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF")
            
            # Color in last row
            cell = ws[f"{get_column_letter(col)}12"]
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")

        # Color in first column
        for row in range(3,12):
            cell = ws[f"{get_column_letter(1)}{row}"]
            cell.fill = data_fill
            cell.font = Font(bold=True, color="FF000000")

        # Add borders between rows
        for col in range(1, len(headers) + 1):
            for row in range(2,13):
                cell = ws[f"{get_column_letter(col)}{row}"]
                cell.border = thin_border

        # Style Alignment
        for col in range(1, len(headers) + 1):
            for row in range(1, len(output) + 3):
                ws[f"{get_column_letter(col)}{row}"].alignment = Alignment(horizontal="center", vertical="center")

        # Autofit columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        # Save file
        wb.save(file_path)  # <-- Save the Decom Dashboard sheet here before overwriting wb
        print("✅ Decom Table is completed!")


        # === Load Excel file and read 'Raw Data' from row 2 ===
        df = pd.read_excel(file_path, sheet_name="Raw Data", header=1)

        # === Clean column names ===
        df.columns = df.columns.str.strip()

        # === Rename relevant columns only from actual headers ===
        df = df.rename(columns={
            "OE Name": "OE",                      # Column A
            "Name.2": "Application",           # Column I
            "Forecast End Date": "ForecastDate",  # Column N
            "Phase": "Phase"                   # Column P
        })

        print("\nRenamed Columns:")
        print(df.columns.tolist())

        # === Convert date and filter only 2025 entries ===
        df["ForecastDate"] = pd.to_datetime(df["ForecastDate"], errors="coerce")
        df = df[df["ForecastDate"].dt.year == 2025]

        # === Auto-generate quarter from forecast date ===
        df["Final_Qtr"] = "Q" + df["ForecastDate"].dt.quarter.astype(str)

        # === Split into In Progress vs Completed ===
        in_progress_df = df[~df["Phase"].isin(["Completed", "Descoped"]) & df["Phase"].notna()]
        completed_df = df[df["Phase"] == "Completed"]

        # === Pivot: In Progress ===
        pivot_in_progress = pd.pivot_table(
            in_progress_df,
            index=["OE", "Application"],
            columns="Final_Qtr",
            aggfunc='size',
            fill_value=0
        )
        pivot_in_progress["Grand Total"] = pivot_in_progress.sum(axis=1)
        pivot_in_progress = pivot_in_progress.reset_index()

        # === Add OE subtotal rows ===
        quarter_cols = [col for col in pivot_in_progress.columns if col.startswith("Q")]
        cols_to_sum = quarter_cols + ["Grand Total"]

        # Create final list of rows
        final_rows = []

        for oe in pivot_in_progress["OE"].unique():
            oe_block = pivot_in_progress[pivot_in_progress["OE"] == oe].copy()
            final_rows.append(oe_block)

            # Create subtotal row
            subtotal = oe_block[cols_to_sum].sum()
            subtotal_row = {
                "OE": f"{oe} Total",
                "Application": ""
            }
            subtotal_row.update(subtotal.to_dict())

            # Append subtotal row as DataFrame
            final_rows.append(pd.DataFrame([subtotal_row]))

        # Combine all into one final DataFrame
        pivot_in_progress_final = pd.concat(final_rows, ignore_index=True)


        # === Blank out duplicate OE names (In Progress)
        previous_oe = None
        for idx, row in pivot_in_progress_final.iterrows():
            current_oe = row["OE"]
            if current_oe == previous_oe:
                pivot_in_progress_final.at[idx, "OE"] = ""
            if not current_oe.endswith("Total") and current_oe != "Grand Total":
                previous_oe = current_oe
            else:
                previous_oe = None  # Reset so new OE group starts showing again



        # === Add final Grand Total row to In Progress Pivot ===
        # Only sum rows where OE ends with 'Total'
        country_totals_ip = pivot_in_progress_final[pivot_in_progress_final["OE"].str.endswith("Total")]
        grand_total_ip = country_totals_ip[cols_to_sum].sum()

        grand_total_row_ip = {
            "OE": "Grand Total",
            "Application": ""
        }
        grand_total_row_ip.update(grand_total_ip.to_dict())

        pivot_in_progress_final = pd.concat(
            [pivot_in_progress_final, pd.DataFrame([grand_total_row_ip])],
            ignore_index=True
        )

        # === Pivot: Completed ===
        pivot_completed = completed_df.groupby(["OE", "Application"]).size().reset_index(name="Total")

        # === Add OE-level Total rows to Completed Pivot ===
        completed_rows = []

        for oe in pivot_completed["OE"].unique():
            block = pivot_completed[pivot_completed["OE"] == oe].copy()
            completed_rows.append(block)

            # Total for OE
            total_val = block["Total"].sum()
            total_row = pd.DataFrame([{
                "OE": f"{oe} Total",
                "Application": "",
                "Total": total_val
            }])
            completed_rows.append(total_row)

        # Combine OE blocks and totals
        pivot_completed_final = pd.concat(completed_rows, ignore_index=True)


        # === Keep only first occurrence of each OE in Completed (before subtotal row)
        # === Blank out duplicate OE names (Completed)
        previous_oe_completed = None
        for idx, row in pivot_completed_final.iterrows():
            current_oe = row["OE"]
            if current_oe == previous_oe_completed:
                pivot_completed_final.at[idx, "OE"] = ""
            if not current_oe.endswith("Total") and current_oe != "Grand Total":
                previous_oe_completed = current_oe
            else:
                previous_oe_completed = None  # Reset

        # === Add Grand Total row at the end ===
        # Filter only rows where OE ends with 'Total' (our country total rows)
        country_totals = pivot_completed_final[pivot_completed_final["OE"].str.endswith("Total")]
        grand_total_value = country_totals["Total"].sum()

        grand_total_row = pd.DataFrame([{
            "OE": "Grand Total",
            "Application": "",
            "Total": grand_total_value
        }])

        pivot_completed_final = pd.concat([pivot_completed_final, grand_total_row], ignore_index=True)

        # === Replace 0s with blanks ===
        pivot_in_progress_final = pivot_in_progress_final.replace(0, "")
        pivot_completed_final = pivot_completed_final.replace(0, "")


        # === Create fresh workbook and worksheet ===
        # === Load existing workbook and add new sheet for pivot tables ===
        wb = load_workbook(file_path)
        if "Pivot Tables" in wb.sheetnames:
            del wb["Pivot Tables"]
        ws = wb.create_sheet("Pivot Tables")

        # === Insert labels at A1 and H1 ===
        ws["A1"] = "2025 Decom - In Progress"
        ws["H1"] = "2025 Decom - Completed"

        bold_font = Font(bold=True)
        # Color only A1 and H1
        label_fill = PatternFill(start_color="D9E1F2", fill_type="solid")
        ws["A1"].fill = label_fill
        ws["A1"].font = bold_font
        ws["H1"].fill = label_fill
        ws["H1"].font = bold_font

        # Optional: center align for better aesthetics
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["H1"].alignment = Alignment(horizontal="center", vertical="center")

        # === Determine start rows for headers ===
        # The data starts at row 4, so headers are also at row 4
        in_progress_header_row = 4
        completed_header_row = 4

        # === Insert In Progress table starting from A4 ===
        # Keep track of the last row of the in-progress table
        last_row_in_progress = 0
        for r_idx, row in enumerate(dataframe_to_rows(pivot_in_progress_final, index=False, header=True), start=in_progress_header_row):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
            last_row_in_progress = r_idx

        # === Insert Completed table starting from H4 ===
        # Keep track of the last row of the completed table
        last_row_completed = 0
        for r_idx, row in enumerate(dataframe_to_rows(pivot_completed_final, index=False, header=True), start=completed_header_row):
            for c_idx, val in enumerate(row, start=8):
                ws.cell(row=r_idx, column=c_idx, value=val)
            last_row_completed = r_idx

        # === Set row height for label and headers ===
        ws.row_dimensions[1].height = 28.35
        ws.row_dimensions[4].height = 28.35

        # === Define styles ===
        header_fill = PatternFill(start_color="DDEBF7", fill_type="solid")
        footer_fill = PatternFill(start_color="DDEBF7", fill_type="solid")
        country_total_fill = PatternFill(start_color="E7E6E6", fill_type="solid")
        bold_font = Font(bold=True)
        no_border = Border()

        # === Style function ===
        def style_sheet(ws, last_row_in_progress, last_row_completed, in_progress_header_row, completed_header_row):
            # Determine the number of columns for each table's data
            num_cols_in_progress = len(pivot_in_progress_final.columns)
            num_cols_completed = len(pivot_completed_final.columns)

            # === Apply styles to In Progress table (A to E or so) ===
            for r in range(1, last_row_in_progress + 1): # Iterate only up to the last row of the in-progress table
                oe_value = ws.cell(row=r, column=1).value  # Column A (In Progress)
                is_header_in_progress = r == in_progress_header_row
                is_grand_total_in_progress = (r == last_row_in_progress) and (oe_value == "Grand Total")
                is_country_total_in_progress = isinstance(oe_value, str) and oe_value.endswith("Total") and oe_value != "Grand Total"

                for c in range(1, num_cols_in_progress + 1):  # Columns A to E (or whatever number of columns the pivot_in_progress_final has)
                    cell = ws.cell(row=r, column=c)
                    cell.alignment = Alignment(vertical="center")  # ✅ NEW LINE
                    align_center = Alignment(horizontal="center", vertical="center")
                    align_left = Alignment(horizontal="left", vertical="center")

                    # Center-align Q3, Q4, Grand Total (adjust if more quarters exist)
                    if cell.column_letter in ["C", "D", "E"]:
                        cell.alignment = align_center
                    elif cell.column == 1:
                        cell.alignment = Alignment(vertical="center", indent=1)
                    else:
                        cell.alignment = align_left

                    if c == 1:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(vertical="center", indent=1)

                    # Exclude B1-E1 from header fill
                    if r == 1 and 1 < c <= num_cols_in_progress: # B1 to E1 are columns 2 to 5 in a 1-indexed range
                        continue # Skip styling for these cells

                    if is_header_in_progress:
                        cell.fill = header_fill
                        cell.font = bold_font
                        cell.border = no_border
                    elif is_grand_total_in_progress or is_country_total_in_progress:
                        cell.fill = footer_fill if is_grand_total_in_progress else country_total_fill
                        cell.font = bold_font

            # === Apply styles to Completed table (H to end) ===
            for r in range(1, last_row_completed + 1): # Iterate only up to the last row of the completed table
                oe_value_right = ws.cell(row=r, column=8).value  # Column H (Completed)
                is_header_completed = r == completed_header_row
                is_grand_total_completed = (r == last_row_completed) and (oe_value_right == "Grand Total")
                is_country_total_completed = isinstance(oe_value_right, str) and oe_value_right.endswith("Total") and oe_value_right != "Grand Total"

                for c in range(8, 8 + num_cols_completed): # Columns H to K (or whatever number of columns the pivot_completed_final has)
                    cell = ws.cell(row=r, column=c)
                    cell.alignment = Alignment(vertical="center")  # ✅ NEW LINE
                    align_center = Alignment(horizontal="center", vertical="center")
                    align_left = Alignment(horizontal="left", vertical="center")

                    # Center-align Total column (last column in completed)
                    if cell.column_letter == get_column_letter(8 + num_cols_completed - 1):  # Last col
                        cell.alignment = align_center
                    elif cell.column == 8:
                        cell.alignment = Alignment(vertical="center", indent=1)
                    else:
                        cell.alignment = align_left

                    if c == 8:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(vertical="center", indent=1)

                    # H1 is already styled, no need to exclude here
                    if is_header_completed:
                        cell.fill = header_fill
                        cell.font = bold_font
                        cell.border = no_border
                    elif is_grand_total_completed or is_country_total_completed:
                        cell.fill = footer_fill if is_grand_total_completed else country_total_fill
                        cell.font = bold_font

        # === Apply styling and autofit ===


        style_sheet(ws, last_row_in_progress, last_row_completed, in_progress_header_row, completed_header_row)

        def auto_adjust_column_width(ws):
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 2

            ws.column_dimensions["F"].width = 2
            ws.column_dimensions["G"].width = 2


        auto_adjust_column_width(ws)

        # === Save final styled file ===
        wb.save(file_path)
        print('yay it works')
    except Exception as e:
        print(f"❌ Something went wrong: {e}")

if __name__ == "__main__":
    main()
